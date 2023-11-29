import os
from datetime import date
import pandas as pd
from openpyxl import load_workbook


class sheetGenerator:
    def __init__(self, excel_path):
        self.excel_path = excel_path

    def getSheetPath(self):
        from openpyxl import load_workbook
        wb2 = load_workbook(self.excel_path)
        current_month = self.getCurrentMonth()
        existed_sheets = self.getSheetList(self.excel_path)
        check_existed = self.checkExistedSheet(existed_sheets, current_month)
        if check_existed == True:
            return self.excel_path, current_month

        wb2.create_sheet(current_month)
        sheet = wb2[current_month]
        header = [['Description', 'path', 'flag']]
        for y in range(len(header)):
            sheet.insert_rows(1, 1)
            for x in range(len(header[y])):
                sheet.cell(row=1 + y,
                           column=2 + x,
                           value=header[y][x])
        wb2.save(self.excel_path)
        wb2.close()
        return self.excel_path, current_month

    @staticmethod
    def getSheetList(excel_path):
        df = pd.read_excel(excel_path, None)
        sheet_list = (pd.ExcelFile(excel_path)).sheet_names
        return sheet_list

    @staticmethod
    def checkExistedSheet(original_list, check_name):
        if check_name not in original_list:
            return False
        return True

    @staticmethod
    def getCurrentMonth():
        this_month = date.today()
        current_month = this_month.strftime("%m")
        return 'T' + current_month

# s = sheetGenerator(r'D:\VTP\python_workspaces\qr_pr\media\24-11-2023.xlsx')
# print(s.getSheetPath()[-1])
